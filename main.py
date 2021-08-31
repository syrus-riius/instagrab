import re
import os
import time
import selenium.webdriver as webdriver
from selenium import webdriver
import urllib.request
from bs4 import BeautifulSoup
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException
import datetime
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws_count = 0

path = "C:\chromedriver_win32\chromedriver.exe"
driver = webdriver.Chrome(path)


#### 저장경로
outpath = "C:/Users/LBuser/Documents/instacrawl2020/"


########################################## defS ############################################

def img_save (outpath, date, title, count, img_src):

    try:
        if not (os.path.isdir(outpath+insta_id+"/")):
            os.makedirs(os.path.join(outpath+insta_id+"/"))
    except OSError as e:
        if e.errno != errno.EEXIST:
            print("Failed to create directory!!!!!")

    imgname = outpath + insta_id + "/" + str(date) + "_" + re.sub('[-=#/?:$}]', '', title) + "_" + insta_id + "_" + str(count) + ".jpg"
    urllib.request.urlretrieve(img_src, imgname)
    print(imgname)
    print(img_src)



# 문서작성
def text_scrap (outpath, date, title, img_all):

    #첫줄을 띄워서 작성하는 경우에 span으로 들어가고
    #그 외 아닌 경우는 h1으로 들어가기 때문에 두 경우를 분리함.
    #text = driver.find_element_by_css_selector('div.EtaWk h1').text


    # 게시물 작성 유형 파악. 한줄 줄바꿈 하면 span으로
    try:
        driver.find_element_by_css_selector('div.EtaWk h1')
        print("Contents is H1.")
        content = driver.find_element_by_css_selector('div.EtaWk h1').text


    except (NoSuchElementException, IndexError):

        try:
            driver.find_element_by_css_selector('div.EtaWk span')
            print("Element is span.")
            content = driver.find_element_by_css_selector('ul.XQXOT div.C4VMK > span').text

        except (NoSuchElementException, IndexError):
            print("내용 없음.")
            content = "(피드 내용 없음.)"

    ### 후.. 너넨 이런거 하지마라..
    ### BeautifulSoup에서 <br>태그 줄바꿈 안된다고 개고생하지마라..
    ### 셀레늄이 짱이다.. 셀레늄 키고 driver.find~~~(element).text 하면 줄바꿈 그대로 가져온다..
    ### 후... 성공했다 드디어..

    # reply_scrap



    #댓글 전부 로드하기
    reply_count = 0


    #댓글 부르기
    # 댓글이 일정 이상 숨겨져있으면 추가버튼 눌러서 다 로드하기

    try:
        while True:
            more_reply = driver.find_element_by_css_selector('ul.XQXOT button.dCJp8')
            driver.execute_script("arguments[0].click()", more_reply)

            print("댓글 추가 버튼 +클릭됨.")
            #reply_count = len(driver.find_elements_by_css_selector('div.EtaWk > ul.XQXOT > ul.Mr508'))
            print("댓글 로드중... (Loading comments...)" )
            time.sleep(0.7)


    except (NoSuchElementException, IndexError):
        pass

        reply_count =len(driver.find_elements_by_css_selector('div.EtaWk > ul.XQXOT > ul.Mr508'))
        time.sleep(1)
        print("댓글 로드 완료! (Completed loading comments!) : 총 " + str(reply_count) + " 개")

    reply_list = []
    i = 0

    # 댓글 순차적 처리
    for i in range(0, reply_count):
        repl = driver.find_elements_by_css_selector('div.EtaWk > ul.XQXOT > ul.Mr508')[i]
        replyer = repl.find_element_by_css_selector('h3').text
        reply = repl.find_element_by_css_selector('span').text

        reply_full = replyer + " : " + reply
        print(reply_full)

        reply_list.append(reply_full)

        # 대댓글 유무에 따라 처리하는 부분
        # 대댓글 오픈하면 div.MGdpg 생성됨

        try:
            show_sub_repl = repl.find_element_by_css_selector('ul.TCSYW')
            print('해당 댓글에 대한 대댓글 발견.')
            coc_count = 0

            coc = show_sub_repl.find_element_by_css_selector('span.EizgU')
            driver.execute_script("arguments[0].click()", coc)
            print("대댓글 추가 버튼 클릭됨.")

            coc_count = len(show_sub_repl.find_elements_by_css_selector('div.ZyFrc'))

            for c in range(0, coc_count):
                sub_reply_select = show_sub_repl.find_elements_by_css_selector('div.ZyFrc')[c]


                # reply_count = len(driver.find_elements_by_css_selector('div.EtaWk > ul.XQXOT > ul.Mr508'))
                print("대댓글 로드중... (Loading comments of comment...)")
                time.sleep(0.5)

                sub_replyer = sub_reply_select.find_element_by_css_selector('h3').text
                sub_reply = sub_reply_select.find_element_by_css_selector('span').text

                sub_reply_full = ("ㄴ " + sub_replyer + " : " + sub_reply)
                print(sub_reply_full)
                reply_list.append(sub_reply_full)



        except (NoSuchElementException, IndexError):
            pass

    reply_all = "\n".join(reply_list)
    img_all_list = "\n".join(img_all)

    now = datetime.datetime.now()
    nowDatetime = now.strftime('%Y-%m-%d %H:%M:%S')

    #파일로 저장
    try:
        if not (os.path.isdir(outpath+insta_id+"/")):
            os.makedirs(os.path.join(outpath+insta_id+"/"))
    except OSError as e:
        if e.errno != errno.EEXIST:
            print("Failed to create directory!!!!!")

    f = open(outpath + insta_id + "/" + str(date) + "_" + re.sub('[-=#/?:$}]', '', title + "_" + insta_id) + ".txt", "w", encoding='UTF8')
    f.write("\n" + "@"+ insta_id + "\n" + date + "\n\n" + feed_url + "\n\n" + img_all_list + "\n\n" + content + "\n\n\n" + reply_all + "\n\nCopyright @" + insta_id + "\n\n\nAmitaProject2020 _ InstaCrawler : " + nowDatetime)
    f.close()
    time.sleep(1)

    global ws_count
    ws_count = ws_count + 1

    ws['A' + str(ws_count)] = str(date)
    ws['B' + str(ws_count)] = str(title)
    ws['C' + str(ws_count)] = str(feed_url)
    ws['D' + str(ws_count)] = str(content)
    wb.save(outpath + insta_id + '.xlsx')





'''
driver.get('https://www.instagram.com')
time.sleep(2)
#############LOGIN##################

id = "righteousnexx"
pw = "fpahsdpdlem3!"

driver.find_element_by_name("username").send_keys(id)
time.sleep(2)
driver.find_element_by_name("password").send_keys(pw)

driver.find_element_by_xpath("//button[@type='submit']").click()
time.sleep(5)
'''
############## Chrome 실행 ##############

insta_url = "https://www.instagram.com/korea_nightview/"


driver.get(insta_url)
insta_id = insta_url[26:]
insta_id = insta_id.replace("/","")


#https://www.instagram.com/korea_nightview/
#driver.get('https://www.instagram.com/ssssi01/')
#driver.get('https://www.instagram.com/mackay_official/')
#driver.get('https://www.instagram.com/righteousnexx/')
#driver.get('https://www.instagram.com/gwangjin_nolja/')
#driver.get('https://www.instagram.com/charles_leclerc/')

time.sleep(2)

driver.find_element_by_css_selector('div.v1Nh3.kIKUG._bz0w').click()

time.sleep(3)



#### 게시물 순환하여 하나씩 저장하기
for i in range(0, 10):
    print(str(i+1) + " 번째 피드 로드 성공")
    time.sleep(1)


    # 전체 페이지 소스 파싱
    html = driver.page_source
    soup = BeautifulSoup(html, 'lxml')


    # URL 주소 추출
    feed_url = driver.current_url
    print(feed_url)

    time.sleep(1)

    try:
        # timestamp
        date_tag = soup.select_one('a.c-Yi7 time')
        date = date_tag.get("title")
        print(date)

    except(AttributeError):
        print("인스타 로그인 오류로 재시작을 추천합니다.")

        driver.get(insta_url)
        print(str(i + 1) + " 번째 피드 로드 성공")
        time.sleep(1)

        # 전체 페이지 소스 파싱
        html = driver.page_source
        soup = BeautifulSoup(html, 'lxml')

        # URL 주소 추출
        feed_url = driver.current_url
        print(feed_url)

        time.sleep(1)

    # 이미지+텍스트 제목 추출용
    #txt_b4 = soup.select_one('ul.XQXOT div.C4VMK span')
    #txt_b4 = soup.select_one('div.EtaWk h1')

    try:
        driver.find_element_by_css_selector('div.EtaWk h1')
        print("Contents is H1.")
        txt_b4 = soup.select_one('div.EtaWk h1')


    except (NoSuchElementException, IndexError):
        try:
            driver.find_element_by_css_selector('div.EtaWk span')
            print("Element is span.")
            txt_b4 = soup.select_one('ul.XQXOT div.C4VMK > span')

        except (NoSuchElementException, IndexError):
            print("내용 없음.")
            txt_b4 = "(내용이 없는 피드)"


    try:
        title = txt_b4.getText()
        title = title[:25]
        print("Feed title is :" + title)
    except:
        title = txt_b4
        print("Feed title is (내용이 없는 피드)")

    count = 0
    img_all = []

    try:  # 여러장의 사진이 있는 경우를 체크하는 예외처리

        #사진이 여러장/한 장 구분은 dir.rQDP3 유무로 확인. div.rQDP3이 없을 경우 except로 넘어감

        driver.find_element_by_css_selector('div.rQDP3')
        print("MultiImages.")


        #인스타의 경우 첫번째/마지막 사진에서는 li가 2개로 존재하고
        # 그 뒤로는 이전/다음 이미지와 함께 3개의 li가 생성됨.

        #따라서 첫번째 사진은 첫번째 li 처리를 따로 해주고
        #이후부터는 두번째 li 처리를 해줘서 끝까지 따올 수 있게 하기 위함.

        #multi_li_first = driver.find_elements_by_css_selector('ul.vi798 li.Ckrof')

        img = soup.select_one('ul.vi798 li.Ckrof div.KL4Bh img')
        img_src = img.get("src")
        count += 1
        #파일 저장

        img_all.append(img_src)
        print(img_all)
        img_save(outpath, date, title, count, img_src)


        # 같은 피드 다음사진
        next_pic = driver.find_element_by_css_selector('button._6CZji')
        next_pic.click()

        try:

            while True: #마지막 사진에 도달해서 예외가 뜰 때까지 무한반복
                html = driver.page_source
                soup = BeautifulSoup(html, 'lxml')
                # multi_li = driver.find_elements_by_css_selector('ul.vi798 li.Ckrof')

                time.sleep(1)

                # 사진이 여러장일 경우 li 형태로 출력. 첫번째 사진은 [0] 이므로,
                # 끝까지는 [1] 경로를 가져오면 된다.
                img = soup.select('ul.vi798 li.Ckrof div.KL4Bh img')[1]
                img_src = img.get("src")

                # 파일 저장
                img_all.append(img_src)
                print(img_all)

                count += 1
                img_save(outpath, date, title, count, img_src)

                # 같은 피드 다음사진
                next_pic = driver.find_element_by_css_selector('button._6CZji')
                next_pic.click()
                time.sleep(0.5)

        except (StaleElementReferenceException, IndexError):
            print("NO MORE IMAGES.")
            text_scrap(outpath, date, title, img_all)
            time.sleep(1)

    except (NoSuchElementException, IndexError):
        ############# 사진이 1장일 경우 ##############

        print("single image")
        # img_src scrap
        img = soup.select_one('div._97aPb div.KL4Bh img')
        img_src = img.get("src")

        # 파일 저장
        count += 1
        img_save(outpath, date, title, count, img_src)
        img_all.append(img_src)
        print(img_all)

        text_scrap(outpath, date, title, img_all)
        time.sleep(1)

    #다음 피드 클릭.
    try:
        driver.find_element_by_css_selector('a._65Bje.coreSpriteRightPaginationArrow').click()
    except (StaleElementReferenceException, IndexError):
        print("다음 피드가 없습니다.")


'''

. if로 element check

if not element:
    print("No Element found")  
else:
    driver.find_element_by_xpath("""Copy XPath""").click()  


body = driver.find_element_by_tag_name('body')
# body를 스크롤하기 위해 tagname이 body로 되어있는것을 추출합니다.
num_of_pagedowns = 50
# 페이지 다운을 몇 번 할지 정해줍니다.(변수로 지정하지 않고 바로 숫자를 입력해도 상관 없습니다.)
while num_of_pagedowns:
  body.send_keys(Keys.PAGE_DOWN)
  # Selenium이 페이지 다운을 할 수 있도록 코드를 입력합니다.
  time.sleep(3.5)
  # 얼마의 시간 뒤에 다시 페이지를 내릴지 시간을 정합니다. 스크롤을 내린 페이지가 로드 될 수 있도록 시간차를 줍니다.
  num_of_pagedowns -= 1

'''